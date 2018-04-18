import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
import subprocess
from openpyxl import load_workbook, Workbook
import Tkinter, Tkconstants, tkFileDialog
import os, sys
from collections import OrderedDict
files = tkFileDialog.askopenfilenames(parent=Tkinter.Tk(),title='Choose a file')
lst = list(files)
noc = 150 #***Number OF Cycles***
CDC= 'CompareDischargeCapacities.xlsx'
full_data={}
full_data["Cycle Number"]=range(1,noc+1)
full_data_ordered=OrderedDict(full_data)
for i in lst:
    file_name=str(i)
    if "xlsx" in file_name:
        wb = load_workbook(filename = i)
        ws = wb['Cycle']
        a,b=os.path.split(i)
        discharge_capacities = []
        battery_name = b[:-5]
        for row in range(3,noc+3): #First two rows irrelevant
            d = {i:ws.cell(row, 5).value} #Discharge capacity column
            e = str(d)
            f = e[:-1]
            g = f.split(".xlsx': ",1)[1] #get discharge capacity value
            discharge_capacities.append(float(g))
        full_data_ordered[battery_name]=discharge_capacities
        discharge_capacities = {}
    else:
        a,b=os.path.split(i)
        battery_name = b[:-4] #Delete.txt
        discharge_capacities = []
        file = open(i,'r')
        values=file.read()
        values = values.split("\t")
        header_location=values.index("RCap_DChg") #Find start of data
        del values[:header_location] #Delete everything before data
        del values[:8] #Start at first discharge capacity value
        data_list = values[0::8] #Get remaining discharge capacity values which occur every 8th item in values
        for item in data_list[0:noc]:
            discharge_capacities.append(float(item))
        full_data_ordered[battery_name]=discharge_capacities
df = pd.DataFrame.from_dict(full_data_ordered)
writer=ExcelWriter(CDC)
df.to_excel(writer,'Sheet1',index=False)
writer.save()
