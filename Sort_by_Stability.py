import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
import subprocess
from openpyxl import load_workbook, Workbook
import Tkinter, Tkconstants, tkFileDialog
import os, sys
from collections import OrderedDict
from operator import itemgetter
files = tkFileDialog.askopenfilenames(parent=Tkinter.Tk(),title='Choose a file')
lst = list(files)
noc = 150 #Number of Cycles
stability_rank= 'sort_by_stability.xlsx'
data={}
full_data=OrderedDict({})
for i in lst:
    file_name=str(i)
    if "xlsx" in file_name:
        wb = load_workbook(filename = i)
        ws = wb['Cycle']
        a,b=os.path.split(i)
        discharge_capacities = []
        discharge_capacities_float = []
        battery_name = b[:-5]
        for row in range(3,noc+3): #First two rows irrelevant
            d = {i:ws.cell(row, 5).value} #discharge capacity column
            e = str(d)
            f = e[:-1]
            g = f.split(".xlsx': ",1)[1] #get discharge capacity value
            discharge_capacities.append(g)
        for item in discharge_capacities:
            discharge_capacities_float.append(float(item))
        dec=[np.diff(discharge_capacities_float)]
        avg=-1*np.mean(dec) #Average Decrease in Capacity Per Cycle
        d = {i:ws.cell(3, 5).value} #First Discharge Capacity Value Location
        e = str(d)
        f = e[:-1]
        g = f.split(".xlsx': ",1)[1]
        fcdc=float(g) #First cycle discharge capacity value
        h = {i:ws.cell(3, 6).value} #First Cycle Coloumbic Efficiency Location
        j = str(h)
        k = j[:-1]
        l = k.split(".xlsx': ",1)[1]
        fcce = float(l) #First cycle efficiency
        avgp = 100*avg/fcdc #Average Percent of Initial Capacity Lost Per Cycle
        allvalues = [avgp, fcdc, fcce]
        data[battery_name]=allvalues
    else:
        a,b=os.path.split(i)
        battery_name = b[:-4] #Delete.txt
        discharge_capacities = []
        discharge_capacities_float = []
        allvalues = []
        file = open(i,'r')
        values=file.read()
        values = values.split("\t")
        header_location=values.index("RCap_DChg") #Find start of data
        del values[:header_location] #Delete everything before data
        del values[:8] #Start at first discharge capacity value
        data_list = values[0::8] #Get remaining discharge capacity values which occur every 8th item in values
        for item in data_list[0:noc]:
            discharge_capacities.append(item)
        for item in discharge_capacities:
            discharge_capacities_float.append(float(item))
        dec=[np.diff(discharge_capacities_float)]
        avg = -1*np.mean(dec) #Average Decrease in Capacity Per Cycle
        fcdc = float(data_list[0]) #First Cycle Discharge Capacity
        fcce = float(values[1]) #First Cycle Coulombic Efficiency
        avgp = 100*avg*(fcdc**-1) #Average Percent of Initial Capacity Lost Per Cycle
        allvalues = [avgp, fcdc, fcce]
        data[battery_name]=allvalues
data_sorted=sorted(data.items(), key=itemgetter(1))
full_data.update(data_sorted)
df = pd.DataFrame.from_dict(full_data).T
writer=ExcelWriter(stability_rank)
df.to_excel(writer,'Sheet1', header=["Percent of Initial Capacity Lost Per Cycle", "First Cycle Discharge Capacity (mAh/g)", "First Cycle Coulombic Efficiency"])
writer.save()
